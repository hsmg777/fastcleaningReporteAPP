from flask_smorest import Blueprint, abort
from flask.views import MethodView
from flask import send_file
from db import db
from models.ReporteMensual import ReporteMensual
from models.Orden import Orden
from models.Gastos import Gastos
from schemas.ReporteMensualSchema import ReporteMensualSchema
from sqlalchemy.sql import text  # Asegúrate de importar esto
import openpyxl
from io import BytesIO

# Crear el Blueprint
blp = Blueprint("ReporteMensual", __name__, url_prefix="/tasks/reportes", description="Operaciones CRUD para Reportes Mensuales")

# Schemas
reporte_schema = ReporteMensualSchema()
reportes_schema = ReporteMensualSchema(many=True)


# Rutas para /tasks/reportes
@blp.route('/')
class ReporteMensualList(MethodView):
    @blp.response(200, ReporteMensualSchema(many=True))  # Define el esquema para la respuesta
    def get(self):
        """Obtener todos los reportes mensuales"""
        reportes = ReporteMensual.query.all()
        return reportes

    @blp.arguments(ReporteMensualSchema)  # Valida la entrada con el esquema
    @blp.response(201, ReporteMensualSchema)  # Define el esquema para la respuesta
    def post(self, data):
        """Crear un nuevo reporte mensual"""
        try:
            nuevo_reporte = ReporteMensual(
                nombreMes=data['nombreMes']
            )
            db.session.add(nuevo_reporte)
            db.session.commit()
            return nuevo_reporte
        except Exception as e:
            abort(400, message=f"Error al crear el reporte mensual: {str(e)}")


# Rutas para /tasks/reportes/<int:id_reporteMensual>
@blp.route('/<int:id_reporteMensual>')
class ReporteMensualResource(MethodView):
    @blp.response(200, ReporteMensualSchema)  # Define el esquema para la respuesta
    def get(self, id_reporteMensual):
        """Obtener un reporte mensual por su ID"""
        reporte = ReporteMensual.query.get_or_404(id_reporteMensual)
        return reporte

    @blp.arguments(ReporteMensualSchema)  # Valida la entrada con el esquema
    @blp.response(200, ReporteMensualSchema)  # Define el esquema para la respuesta
    def put(self, data, id_reporteMensual):
        """Actualizar un reporte mensual existente"""
        reporte = ReporteMensual.query.get_or_404(id_reporteMensual)
        if 'nombreMes' in data:
            reporte.nombreMes = data['nombreMes']
        db.session.commit()
        return reporte

    @blp.response(204)  # Sin contenido en la respuesta
    def delete(self, id_reporteMensual):
        """Eliminar un reporte mensual"""
        reporte = ReporteMensual.query.get_or_404(id_reporteMensual)
        db.session.delete(reporte)
        db.session.commit()
        return '', 204


# Nuevo endpoint para exportar datos a Excel
@blp.route('/exportar/<int:id_reporteMensual>', methods=['GET'])
def exportar_reporte(id_reporteMensual):
    """Exportar reporte mensual, órdenes y gastos a Excel"""
    try:
        # Consultar el reporte
        reporte = ReporteMensual.query.get_or_404(id_reporteMensual)

        # Consultar las órdenes asociadas al reporte
        ordenes = db.session.execute(
            text("SELECT numeroOrden AS orden, valor, fecha FROM Orden WHERE id_reporteMensual = :id_reporteMensual"),
            {'id_reporteMensual': id_reporteMensual}
        ).fetchall()

        # Consultar los gastos asociados al reporte
        gastos = db.session.execute(
            text("SELECT nombreGasto AS nombre, valor, fecha FROM Gastos WHERE id_reporteMensual = :id_reporteMensual"),
            {'id_reporteMensual': id_reporteMensual}
        ).fetchall()

        # Crear el archivo Excel
        output = BytesIO()
        workbook = openpyxl.Workbook()
        
        # Agregar hoja para el resumen del reporte
        sheet = workbook.active
        sheet.title = "Resumen"
        sheet.append(["Mes", "Total Neto", "Total Gastos", "Ganancia"])
        sheet.append([reporte.nombreMes, reporte.totalNeto, reporte.totalGastos, reporte.ganancia])

        # Agregar hoja para órdenes
        ordenes_sheet = workbook.create_sheet(title="Órdenes")
        ordenes_sheet.append(["Orden", "Valor", "Fecha"])
        for orden in ordenes:
            ordenes_sheet.append(list(orden))

        # Agregar hoja para gastos
        gastos_sheet = workbook.create_sheet(title="Gastos")
        gastos_sheet.append(["Nombre del Gasto", "Valor", "Fecha"])
        for gasto in gastos:
            gastos_sheet.append(list(gasto))

        # Guardar el archivo Excel en memoria
        workbook.save(output)
        output.seek(0)

        # Enviar el archivo como respuesta
        return send_file(
            output,
            as_attachment=True,
            download_name=f"reporte_{reporte.nombreMes}.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    except Exception as e:
        abort(400, message=f"Error al exportar el reporte: {str(e)}")
